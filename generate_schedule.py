#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
节目排期表每日图片生成工具 - 动态列检测版
新增功能：
1. 自动读取 Excel 第2行数据作为节目列名称，不再依赖写死的 (列索引, 名称)。
2. 增加 `START_COL_INDEX` 配置项，从指定的列开始扫描（目前设定为48列）。
"""
import os
import sys
import re
import argparse
from datetime import datetime, timedelta
from collections import defaultdict
import openpyxl
from PIL import Image, ImageDraw, ImageFont

# ==================== 配置区 ====================
EXCEL_PATH = "节目排期表.xlsx"
OUTPUT_DIR = "排期图片输出"

# 核心更新：指定从 Excel 的哪一列开始读取节目表头
# 根据你的表格，第 48 列之前是旧表头，48 以后是新排期
START_COL_INDEX = 48

# 图片尺寸
IMAGE_WIDTH = 1200
ROW_HEIGHT = 80
HEADER_HEIGHT = 60
MARGIN = 20
COL_WIDTH_RATIO = [2.5, 5.5, 1.5]

# 全局颜色
HEADER_BG = "#D0E8E8"
HEADER_TEXT = "#000000"
BORDER_COLOR = "#B0B0B0"
CONTENT_BG = "#FFFFFF"
PROGRAM_NAME_TEXT = "#000000"
ACTIVITY_TEXT = "#C00000"
ACTIVITY_CONTENT_BG = "#FFFF00"
ACTIVITY_CONTENT_FONT = "#FF0000"

# Office默认主题色转换
DEFAULT_THEME_COLORS = [
    '000000','FFFFFF','44546A','E7E6E6','4472C4','ED7D31','A5A5A5','FFC000','5B9BD5','70AD47'
]

# ==================== 颜色转换工具 ====================
def color_to_rgb(color_obj):
    if not color_obj:
        return '#FFFFFF'
    if color_obj.type == 'rgb':
        rgb_raw = color_obj.rgb
        if not isinstance(rgb_raw, str):
            return '#FFFFFF'
        rgb = str(rgb_raw).upper()
        if len(rgb) == 8:
            hex_str = rgb[2:]
        elif len(rgb) == 6:
            hex_str = rgb
        else:
            return '#FFFFFF'
        if not re.fullmatch(r'[0-9A-F]{6}', hex_str):
            return '#FFFFFF'
        try:
            r = int(hex_str[0:2], 16)
            g = int(hex_str[2:4], 16)
            b = int(hex_str[4:6], 16)
        except ValueError:
            return '#FFFFFF'
        r = max(0, min(255, r))
        g = max(0, min(255, g))
        b = max(0, min(255, b))
        return f'#{r:02X}{g:02X}{b:02X}'
    elif color_obj.type == 'theme':
        tidx = color_obj.theme
        tint = color_obj.tint or 0.0
        if 0 <= tidx < len(DEFAULT_THEME_COLORS):
            base = DEFAULT_THEME_COLORS[tidx]
            try:
                r = int(base[0:2],16)
                g = int(base[2:4],16)
                b = int(base[4:6],16)
            except ValueError:
                return '#FFFFFF'
            if tint >= 0:
                r = int(r + (255 - r)*tint)
                g = int(g + (255 - g)*tint)
                b = int(b + (255 - b)*tint)
            else:
                r = int(r*(1+tint))
                g = int(g*(1+tint))
                b = int(b*(1+tint))
            r = max(0, min(255, r))
            g = max(0, min(255, g))
            b = max(0, min(255, b))
            return f'#{r:02X}{g:02X}{b:02X}'
    return '#FFFFFF'

def get_column_bg_color(ws, col):
    cell = ws.cell(row=2, column=col)
    if cell.fill and cell.fill.fill_type and cell.fill.fgColor:
        return color_to_rgb(cell.fill.fgColor)
    return '#FFFFFF'

# ==================== 文本清洗+识别 ====================
def clean_text(s):
    if not s:
        return ""
    s = str(s).strip()
    s = re.sub(r'\s+', '', s)
    return s

def get_content_type(content):
    s = clean_text(content)
    if "谷养正道" in s:
        return "gyzd"
    elif "家有老中医" in s:
        return "jylzy"
    elif "御医有方" in s or "对话大医生" in s:
        return "yuyi"
    else:
        return "common"

# ==================== 基础解析工具 ====================
def get_excel_date(date_obj):
    base = datetime(1899, 12, 30)
    return (date_obj - base).days

def is_activity_cell(cell):
    red_font = False
    yellow_bg = False
    if cell.font and cell.font.color and cell.font.color.rgb:
        cr = str(cell.font.color.rgb).upper()
        if cr in ['FFFF0000','00FF0000','FFC00000','FFE00000']:
            red_font = True
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        br = str(cell.fill.fgColor.rgb).upper()
        if br in ['FFFFFF00','00FFFF00','FFFFE699','FFFFF2CC','FFFFCC00']:
            yellow_bg = True
    return red_font or yellow_bg

def extract_round_and_episode(content):
    s = clean_text(content)
    # 谷养正道X轮Y期
    m1 = re.search(r'谷养正道(\d+)轮(\d+)', s)
    if m1:
        return int(m1.group(1)), int(m1.group(2))
    # 家有老中医X-xxx
    m2 = re.search(r'家有老中医(\d+)-', s)
    if m2:
        return 1, int(m2.group(1))
    # 御医有方X
    m3 = re.search(r'御医有方(\d+)', s)
    if m3:
        return 1, int(m3.group(1))
    # 《对话大医生》X
    m4 = re.search(r'《对话大医生》(\d+)', s)
    if m4:
        return 1, int(m4.group(1))
    # 通用第X期
    m5 = re.search(r'第(\d+)期', s)
    if m5:
        return 1, int(m5.group(1))
    return None, None

# ==================== 全局活动区间缓存 ====================
def analyze_all_program_activity(ws):
    cache = {
        "gyzd": defaultdict(list),
        "jylzy": defaultdict(list),
        "yuyi": defaultdict(list),
        "common": defaultdict(list)
    }
    content_map = {
        "gyzd": defaultdict(dict),
        "jylzy": defaultdict(dict),
        "yuyi": defaultdict(dict),
        "common": defaultdict(dict)
    }
    for row in range(3, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if not val:
                continue
            ctype = get_content_type(val)
            r_num, ep = extract_round_and_episode(val)
            if ep is None:
                continue
            content_map[ctype][r_num][ep] = str(val)
            if is_activity_cell(cell):
                cache[ctype][r_num].append(ep)

    final_range = {"gyzd":{}, "jylzy":{}, "yuyi":{}, "common":{}}
    for prog_type in cache.keys():
        round_eps = cache[prog_type]
        for r in sorted(round_eps.keys()):
            eps = sorted(list(set(round_eps[r])))
            ranges = []
            if not eps:
                continue
            start = eps[0]
            prev = eps[0]
            for ep in eps[1:]:
                if ep - prev > 1:
                    c = content_map[prog_type].get(start, "")
                    if "预热" in c or "报名" in c:
                        start += 1
                    if start <= prev:
                        ranges.append((start, prev))
                    start = ep
                prev = ep
            c = content_map[prog_type].get(start, "")
            if "预热" in c or "报名" in c:
                start += 1
            if start <= prev:
                ranges.append((start, prev))
            final_range[prog_type][r] = ranges
    return final_range

def find_next_segment(current_ep, seg_list):
    if not seg_list:
        return None
    for seg in seg_list:
        s, e = seg
        if s <= current_ep <= e:
            return seg
    for seg in seg_list:
        s, e = seg
        if s > current_ep:
            return seg
    return seg_list[-1] if seg_list else None

# ==================== 数据加载 (核心修改部分) ====================
def load_schedule_data(excel_path):
    if not os.path.exists(excel_path):
        xlsx = [f for f in os.listdir() if f.endswith(".xlsx")]
        if xlsx:
            excel_path = xlsx[0]
        else:
            raise FileNotFoundError("无Excel文件")
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb.active

    # ===================== 新增自动扫描列功能 =====================
    program_config = []
    # 从 START_COL_INDEX 列开始扫描第2行
    for col_idx in range(START_COL_INDEX, ws.max_column + 1):
        header_cell = ws.cell(row=2, column=col_idx)
        col_name = str(header_cell.value).strip() if header_cell.value else ""

        # 如果表头为空，或者表头内容包含“播放/时间”这类关键字（帮你过滤掉前面不要的列），则停止扫描
        if not col_name:
            break
        if "播放" in col_name or "时间" in col_name:
            continue

        # 处理换行符
        col_name = col_name.replace(" ", "\n")
        program_config.append((col_idx, col_name))
    # ==============================================================

    all_prog_range = analyze_all_program_activity(ws)
    col_bg = {}
    for cidx, _ in program_config:
        col_bg[cidx] = get_column_bg_color(ws, cidx)
    return ws, all_prog_range, col_bg, program_config

def find_date_row(ws, target_date):
    target_serial = get_excel_date(target_date)
    for row in range(3, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, (int, float)) and abs(v - target_serial) < 0.5:
            return row
    return None

def format_activity_label(c_type, r_num, seg):
    if not seg:
        return "活动期待定"
    base = f"{seg[0]}-{seg[1]}期活动"
    if c_type == "gyzd" and r_num:
        return f"{r_num}轮{base}"
    else:
        return base

def get_day_schedule(ws, target_date, all_prog_range, column_bg_colors, program_config):
    row = find_date_row(ws, target_date)
    if row is None:
        return None
    schedule = []
    for idx, (col_idx, prog_name) in enumerate(program_config):
        cell = ws.cell(row=row, column=col_idx)
        raw_content = str(cell.value).strip() if cell.value else ""
        content = clean_text(raw_content)
        is_act = is_activity_cell(cell)
        c_type = get_content_type(content)
        r_num, ep = extract_round_and_episode(content)
        bg_color = column_bg_colors.get(col_idx, "#FFFFFF")

        act_label = "活动期待定"
        if c_type and r_num and ep:
            seg_list = all_prog_range[c_type].get(r_num, [])
            seg = find_next_segment(ep, seg_list)
            if not seg and c_type == "gyzd":
                next_r = r_num + 1
                next_seg_list = all_prog_range[c_type].get(next_r, [])
                if next_seg_list:
                    seg = next_seg_list[0]
            act_label = format_activity_label(c_type, r_num if c_type == "gyzd" else 1, seg)

        schedule.append({
            "name": prog_name,
            "content": raw_content,
            "bg_color": bg_color,
            "activity": act_label,
            "is_activity": is_act,
            "episode": ep,
            "round_num": r_num
        })
    return schedule

# ==================== 图片绘制 ====================
def get_font(size, bold=False):
    # 修改为优先读取当前目录下上传的 simhei.ttf
    paths = [
        "./simhei.ttf",                     # 1. 当前目录下的字体文件
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", # 2. PythonAnywhere自带英文字体(兜底)
        "/System/Library/Fonts/PingFang.ttc"
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, size)
            except:
                continue
    return ImageFont.load_default()

def wrap_text(text, font, max_w):
    if not text:
        return [""]
    lines = []
    for para in text.split("\n"):
        line = ""
        for ch in para:
            test = line + ch
            bw = font.getbbox(test)[2] - font.getbbox(test)[0]
            if bw <= max_w:
                line = test
            else:
                lines.append(line)
                line = ch
        if line:
            lines.append(line)
    return lines

def generate_schedule_image(schedule, target_date, out_path):
    total_w = IMAGE_WIDTH - MARGIN * 2
    col_w = [int(total_w * r / sum(COL_WIDTH_RATIO)) for r in COL_WIDTH_RATIO]
    f_content = get_font(16)
    f_name = get_font(18, bold=True)
    f_act = get_font(16)
    f_header = get_font(28, bold=True)
    row_h_list = []
    for item in schedule:
        n_line = len(item["name"].split("\n"))
        h1 = n_line * 28
        c_line = wrap_text(item["content"], f_content, col_w[1]-20)
        h2 = len(c_line)*24
        a_line = wrap_text(item["activity"], f_act, col_w[2]-10)
        h3 = len(a_line)*24
        rh = max(h1,h2,h3)+20
        row_h_list.append(max(rh, ROW_HEIGHT))
    total_h = HEADER_HEIGHT + sum(row_h_list) + MARGIN*2
    img = Image.new("RGB", (IMAGE_WIDTH, total_h), "white")
    draw = ImageDraw.Draw(img)
    # 标题栏
    h_y = MARGIN
    title = f"{target_date.month}月{target_date.day}日排期"
    draw.rectangle([MARGIN, h_y, IMAGE_WIDTH-MARGIN, h_y+HEADER_HEIGHT], fill=HEADER_BG)
    tw = f_header.getbbox(title)[2] - f_header.getbbox(title)[0]
    tx = MARGIN + (total_w - tw)//2
    ty = h_y + (HEADER_HEIGHT - (f_header.getbbox(title)[3]-f_header.getbbox(title)[1]))//2
    draw.text((tx, ty), title, fill=HEADER_TEXT, font=f_header)
    cur_y = h_y + HEADER_HEIGHT
    left_x = MARGIN
    for i, item in enumerate(schedule):
        rh = row_h_list[i]
        # 第一列背景
        draw.rectangle([left_x, cur_y, left_x+col_w[0], cur_y+rh], fill=item["bg_color"])
        name_lines = item["name"].split("\n")
        total_n_h = len(name_lines)*28
        n_start_y = cur_y + (rh - total_n_h)//2
        for j, l in enumerate(name_lines):
            lw = f_name.getbbox(l)[2] - f_name.getbbox(l)[0]
            lx = left_x + (col_w[0] - lw)//2
            ly = n_start_y + j*28
            draw.text((lx, ly), l, fill=PROGRAM_NAME_TEXT, font=f_name)
        # 第二列内容
        c_x = left_x + col_w[0]
        if item["is_activity"]:
            cb = ACTIVITY_CONTENT_BG
            cf = ACTIVITY_CONTENT_FONT
        else:
            cb = CONTENT_BG
            cf = "#000000"
        draw.rectangle([c_x, cur_y, c_x+col_w[1], cur_y+rh], fill=cb)
        c_lines = wrap_text(item["content"], f_content, col_w[1]-20)
        cy = cur_y + 10
        for line in c_lines:
            draw.text((c_x+10, cy), line, fill=cf, font=f_content)
            cy += 24
        # 第三列活动标注
        a_x = c_x + col_w[1]
        draw.rectangle([a_x, cur_y, a_x+col_w[2], cur_y+rh], fill=CONTENT_BG)
        a_lines = wrap_text(item["activity"], f_act, col_w[2]-10)
        total_a_h = len(a_lines)*24
        a_start_y = cur_y + (rh - total_a_h)//2
        af = f_act
        if item["is_activity"]:
            af = get_font(16, bold=True)
        for j, l in enumerate(a_lines):
            aw = af.getbbox(l)[2] - af.getbbox(l)[0]
            ax = a_x + (col_w[2] - aw)//2
            ay = a_start_y + j*24
            draw.text((ax, ay), l, fill=ACTIVITY_TEXT, font=af)
        # 边框
        draw.rectangle([left_x, cur_y, IMAGE_WIDTH-MARGIN, cur_y+rh], outline=BORDER_COLOR, width=1)
        draw.line([c_x, cur_y, c_x, cur_y+rh], fill=BORDER_COLOR, width=1)
        draw.line([a_x, cur_y, a_x, cur_y+rh], fill=BORDER_COLOR, width=1)
        cur_y += rh
    # 外框
    draw.rectangle([MARGIN, h_y, IMAGE_WIDTH-MARGIN, cur_y], outline=BORDER_COLOR, width=1)
    os.makedirs(os.path.dirname(out_path) if os.path.dirname(out_path) else ".", exist_ok=True)
    img.save(out_path, "PNG", quality=95)
    return out_path

# ==================== 主入口 ====================
def main():
    parser = argparse.ArgumentParser(description="排期图生成（修复活动期跳转逻辑）")
    parser.add_argument("-d", "--date", help="指定日期 YYYY-MM-DD")
    parser.add_argument("-e", "--excel", help="Excel路径")
    parser.add_argument("-o", "--output", help="输出图片路径")
    parser.add_argument("-t", "--today", action="store_true", help="生成今日")
    parser.add_argument("-v", "--verbose", action="store_true", help="打印详情")
    args = parser.parse_args()

    if args.date:
        try:
            target_dt = datetime.strptime(args.date, "%Y-%m-%d")
        except ValueError:
            print("日期格式错误，使用 YYYY-MM-DD")
            sys.exit(1)
    elif args.today:
        target_dt = datetime.now()
    else:
        target_dt = datetime.now() + timedelta(days=1)
    target_dt = target_dt.replace(hour=0, minute=0, second=0, microsecond=0)

    excel_p = args.excel or EXCEL_PATH
    print(f"正在生成 {target_dt.month}月{target_dt.day} 排期")
    try:
        ws, all_range, col_bg, program_config = load_schedule_data(excel_p)
    except Exception as e:
        print(f"加载Excel失败：{e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    schedule = get_day_schedule(ws, target_dt, all_range, col_bg, program_config)
    if schedule is None:
        print("未找到对应日期排期数据")
        sys.exit(1)

    if args.output:
        out_p = args.output
    else:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        d_str = target_dt.strftime("%Y%m%d")
        out_p = os.path.join(OUTPUT_DIR, f"节目排期_{d_str}.png")

    try:
        generate_schedule_image(schedule, target_dt, out_p)
        print(f"\n✅ 图片生成完成：{out_p}")
    except Exception as e:
        print(f"绘图失败：{e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    if args.verbose:
        print(f"\n{'='*60}")
        print(f"{target_dt.year}年{target_dt.month}月{target_dt.day} 活动期明细")
        print(f"{'='*60}")
        for item in schedule:
            stat = "🔴当期活动" if item["is_activity"] else "⚪日常档"
            rn = f"第{item['round_num']}轮" if item["round_num"] else ""
            epn = f"第{item['episode']}期" if item["episode"] else "未知期"
            print(f"\n【{item['name'].split(chr(10))[0]}】")
            print(f"状态：{stat} {rn} {epn}")
            print(f"活动标注：{item['activity']}")
        act_cnt = sum(1 for i in schedule if i["is_activity"])
        print(f"\n当期活动节目总数：{act_cnt}/{len(schedule)}")

if __name__ == "__main__":
    main()